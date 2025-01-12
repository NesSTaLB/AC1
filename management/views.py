from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import permission_required
from django.http import HttpResponse
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
import csv
import json
from django.utils.encoding import smart_str
from .models import Customer, Engineer, Technician, Product, Project, Quotation, MaintenanceTask, Resource, MaintenanceSchedule, Report
from .forms import CustomUserCreationForm, CustomAuthenticationForm, ExportColumnsForm, CustomerForm, EngineerForm, TechnicianForm, ProductForm, ProjectForm, QuotationForm, MaintenanceTaskForm, ResourceForm, MaintenanceScheduleForm, ReportForm

# الصفحة الرئيسية
def home(request):
    return render(request, 'management/home.html')

# تسجيل المستخدم الجديد
def register(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect('home')
    else:
        form = CustomUserCreationForm()
    return render(request, 'management/register.html', {'form': form})

# تسجيل الدخول
def user_login(request):
    if request.method == 'POST':
        form = CustomAuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('home')
    else:
        form = CustomAuthenticationForm()
    return render(request, 'management/login.html', {'form': form})

# تسجيل الخروج
def user_logout(request):
    logout(request)
    return redirect('home')

# العملاء
def customer_list(request):
    customers = Customer.objects.all()
    return render(request, 'management/customer_list.html', {'customers': customers})

def add_customer(request):
    if request.method == 'POST':
        form = CustomerForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('customer_list')
    else:
        form = CustomerForm()
    return render(request, 'management/add_customer.html', {'form': form})

def customer_detail(request, customer_id):
    customer = get_object_or_404(Customer, id=customer_id)
    return render(request, 'management/customer_detail.html', {'customer': customer})

def edit_customer(request, customer_id):
    customer = get_object_or_404(Customer, id=customer_id)
    if request.method == 'POST':
        form = CustomerForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()
            return redirect('customer_detail', customer_id=customer.id)
    else:
        form = CustomerForm(instance=customer)
    return render(request, 'management/edit_customer.html', {'form': form, 'customer': customer})

def delete_customer(request, customer_id):
    customer = get_object_or_404(Customer, id=customer_id)
    if request.method == 'POST':
        customer.delete()
        return redirect('customer_list')
    return render(request, 'management/delete_customer.html', {'customer': customer})

# المهندسين
def engineer_list(request):
    engineers = Engineer.objects.all()
    return render(request, 'management/engineer_list.html', {'engineers': engineers})

def add_engineer(request):
    if request.method == 'POST':
        form = EngineerForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('engineer_list')
    else:
        form = EngineerForm()
    return render(request, 'management/add_engineer.html', {'form': form})

def engineer_detail(request, engineer_id):
    engineer = get_object_or_404(Engineer, id=engineer_id)
    return render(request, 'management/engineer_detail.html', {'engineer': engineer})

def edit_engineer(request, engineer_id):
    engineer = get_object_or_404(Engineer, id=engineer_id)
    if request.method == 'POST':
        form = EngineerForm(request.POST, instance=engineer)
        if form.is_valid():
            form.save()
            return redirect('engineer_detail', engineer_id=engineer.id)
    else:
        form = EngineerForm(instance=engineer)
    return render(request, 'management/edit_engineer.html', {'form': form, 'engineer': engineer})

def delete_engineer(request, engineer_id):
    engineer = get_object_or_404(Engineer, id=engineer_id)
    if request.method == 'POST':
        engineer.delete()
        return redirect('engineer_list')
    return render(request, 'management/delete_engineer.html', {'engineer': engineer})

# الفنيين
def technician_list(request):
    technicians = Technician.objects.all()
    return render(request, 'management/technician_list.html', {'technicians': technicians})

def add_technician(request):
    if request.method == 'POST':
        form = TechnicianForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('technician_list')
    else:
        form = TechnicianForm()
    return render(request, 'management/add_technician.html', {'form': form})

def technician_detail(request, technician_id):
    technician = get_object_or_404(Technician, id=technician_id)
    return render(request, 'management/technician_detail.html', {'technician': technician})

def edit_technician(request, technician_id):
    technician = get_object_or_404(Technician, id=technician_id)
    if request.method == 'POST':
        form = TechnicianForm(request.POST, instance=technician)
        if form.is_valid():
            form.save()
            return redirect('technician_detail', technician_id=technician.id)
    else:
        form = TechnicianForm(instance=technician)
    return render(request, 'management/edit_technician.html', {'form': form, 'technician': technician})

def delete_technician(request, technician_id):
    technician = get_object_or_404(Technician, id=technician_id)
    if request.method == 'POST':
        technician.delete()
        return redirect('technician_list')
    return render(request, 'management/delete_technician.html', {'technician': technician})

# المنتجات
def product_list(request):
    products = Product.objects.all()
    return render(request, 'management/product_list.html', {'products': products})

def add_product(request):
    if request.method == 'POST':
        form = ProductForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('product_list')
    else:
        form = ProductForm()
    return render(request, 'management/add_product.html', {'form': form})

def product_detail(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    return render(request, 'management/product_detail.html', {'product': product})

def edit_product(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    if request.method == 'POST':
        form = ProductForm(request.POST, instance=product)
        if form.is_valid():
            form.save()
            return redirect('product_detail', product_id=product.id)
    else:
        form = ProductForm(instance=product)
    return render(request, 'management/edit_product.html', {'form': form, 'product': product})

def delete_product(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    if request.method == 'POST':
        product.delete()
        return redirect('product_list')
    return render(request, 'management/delete_product.html', {'product': product})

# المشاريع
def project_list(request):
    projects = Project.objects.all().select_related('customer', 'engineer__user', 'technician__user')
    return render(request, 'management/project_list.html', {'projects': projects})

def add_project(request):
    if request.method == 'POST':
        form = ProjectForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('project_list')
    else:
        form = ProjectForm()
    return render(request, 'management/add_project.html', {'form': form})

def project_detail(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    return render(request, 'management/project_detail.html', {'project': project})

def edit_project(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    if request.method == 'POST':
        form = ProjectForm(request.POST, request.FILES, instance=project)
        if form.is_valid():
            form.save()
            return redirect('project_detail', project_id=project.id)
    else:
        form = ProjectForm(instance=project)
    return render(request, 'management/edit_project.html', {'form': form, 'project': project})

def delete_project(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    if request.method == 'POST':
        project.delete()
        return redirect('project_list')
    return render(request, 'management/delete_project.html', {'project': project})

# عروض الأسعار
def quotation_list(request):
    quotations = Quotation.objects.all()
    return render(request, 'management/quotation_list.html', {'quotations': quotations})

def add_quotation(request):
    if request.method == 'POST':
        form = QuotationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('quotation_list')
    else:
        form = QuotationForm()
    return render(request, 'management/add_quotation.html', {'form': form})

def quotation_detail(request, quotation_id):
    quotation = get_object_or_404(Quotation, id=quotation_id)
    return render(request, 'management/quotation_detail.html', {'quotation': quotation})

def edit_quotation(request, quotation_id):
    quotation = get_object_or_404(Quotation, id=quotation_id)
    if request.method == 'POST':
        form = QuotationForm(request.POST, instance=quotation)
        if form.is_valid():
            form.save()
            return redirect('quotation_detail', quotation_id=quotation.id)
    else:
        form = QuotationForm(instance=quotation)
    return render(request, 'management/edit_quotation.html', {'form': form, 'quotation': quotation})

def delete_quotation(request, quotation_id):
    quotation = get_object_or_404(Quotation, id=quotation_id)
    if request.method == 'POST':
        quotation.delete()
        return redirect('quotation_list')
    return render(request, 'management/delete_quotation.html', {'quotation': quotation})

# مهام الصيانة
def maintenance_task_list(request):
    maintenance_tasks = MaintenanceTask.objects.all().select_related('technician__user')
    return render(request, 'management/maintenance_task_list.html', {'maintenance_tasks': maintenance_tasks})

def add_maintenance_task(request):
    if request.method == 'POST':
        form = MaintenanceTaskForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('maintenance_task_list')
    else:
        form = MaintenanceTaskForm()
    return render(request, 'management/add_maintenance_task.html', {'form': form})

def maintenance_task_detail(request, task_id):
    task = get_object_or_404(MaintenanceTask, id=task_id)
    return render(request, 'management/maintenance_task_detail.html', {'task': task})

def edit_maintenance_task(request, task_id):
    task = get_object_or_404(MaintenanceTask, id=task_id)
    if request.method == 'POST':
        form = MaintenanceTaskForm(request.POST, instance=task)
        if form.is_valid():
            form.save()
            return redirect('maintenance_task_detail', task_id=task.id)
    else:
        form = MaintenanceTaskForm(instance=task)
    return render(request, 'management/edit_maintenance_task.html', {'form': form, 'task': task})

def delete_maintenance_task(request, task_id):
    task = get_object_or_404(MaintenanceTask, id=task_id)
    if request.method == 'POST':
        task.delete()
        return redirect('maintenance_task_list')
    return render(request, 'management/delete_maintenance_task.html', {'task': task})

# الموارد
def resource_list(request):
    resources = Resource.objects.all()
    return render(request, 'management/resource_list.html', {'resources': resources})

def add_resource(request):
    if request.method == 'POST':
        form = ResourceForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('resource_list')
    else:
        form = ResourceForm()
    return render(request, 'management/add_resource.html', {'form': form})

def edit_resource(request, resource_id):
    resource = get_object_or_404(Resource, id=resource_id)
    if request.method == 'POST':
        form = ResourceForm(request.POST, instance=resource)
        if form.is_valid():
            form.save()
            return redirect('resource_list')
    else:
        form = ResourceForm(instance=resource)
    return render(request, 'management/edit_resource.html', {'form': form, 'resource': resource})

def delete_resource(request, resource_id):
    resource = get_object_or_404(Resource, id=resource_id)
    if request.method == 'POST':
        resource.delete()
        return redirect('resource_list')
    return render(request, 'management/delete_resource.html', {'resource': resource})

# جدولة الصيانة
def maintenance_schedule_list(request):
    schedules = MaintenanceSchedule.objects.all()
    return render(request, 'management/maintenance_schedule_list.html', {'schedules': schedules})

def add_maintenance_schedule(request):
    if request.method == 'POST':
        form = MaintenanceScheduleForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('maintenance_schedule_list')
    else:
        form = MaintenanceScheduleForm()
    return render(request, 'management/add_maintenance_schedule.html', {'form': form})

def edit_maintenance_schedule(request, schedule_id):
    schedule = get_object_or_404(MaintenanceSchedule, id=schedule_id)
    if request.method == 'POST':
        form = MaintenanceScheduleForm(request.POST, instance=schedule)
        if form.is_valid():
            form.save()
            return redirect('maintenance_schedule_list')
    else:
        form = MaintenanceScheduleForm(instance=schedule)
    return render(request, 'management/edit_maintenance_schedule.html', {'form': form, 'schedule': schedule})

def delete_maintenance_schedule(request, schedule_id):
    schedule = get_object_or_404(MaintenanceSchedule, id=schedule_id)
    if request.method == 'POST':
        schedule.delete()
        return redirect('maintenance_schedule_list')
    return render(request, 'management/delete_maintenance_schedule.html', {'schedule': schedule})

# تصدير البيانات
def export_customers_excel(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="customers.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"

    # إضافة العناوين
    ws['A1'] = 'اسم العميل'
    ws['B1'] = 'الهاتف'
    ws['C1'] = 'البريد الإلكتروني'
    ws['D1'] = 'العنوان'

    # إضافة البيانات
    customers = Customer.objects.all()
    for idx, customer in enumerate(customers, start=2):
        ws[f'A{idx}'] = customer.name
        ws[f'B{idx}'] = customer.phone
        ws[f'C{idx}'] = customer.email
        ws[f'D{idx}'] = customer.address

    wb.save(response)
    return response

def export_customers_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="customers.pdf"'

    p = canvas.Canvas(response)
    p.drawString(100, 750, "قائمة العملاء")

    customers = Customer.objects.all()
    y = 730
    for customer in customers:
        p.drawString(100, y, f"اسم العميل: {customer.name}")
        p.drawString(100, y - 20, f"الهاتف: {customer.phone}")
        p.drawString(100, y - 40, f"البريد الإلكتروني: {customer.email}")
        p.drawString(100, y - 60, f"العنوان: {customer.address}")
        y -= 100

    p.showPage()
    p.save()
    return response

# نظام التقارير المتقدم
def create_report(request):
    if request.method == 'POST':
        form = ReportForm(request.POST)
        columns_form = ExportColumnsForm(request.POST)
        if form.is_valid() and columns_form.is_valid():
            report = form.save(commit=False)
            report.created_by = request.user
            
            # تحميل الفلاتر من النموذج
            filters = json.loads(request.POST.get('filters', '{}'))
            
            # تطبيق الفلاتر على المشاريع
            projects = Project.objects.all()
            if 'start_date' in filters and filters['start_date']:
                projects = projects.filter(start_date__gte=filters['start_date'])
            if 'end_date' in filters and filters['end_date']:
                projects = projects.filter(end_date__lte=filters['end_date'])
            if 'status' in filters and filters['status'] != 'all':
                projects = projects.filter(status=filters['status'])
            
            # حفظ التقرير
            report.filters = filters
            report.save()
            
            # تحديد الأعمدة المطلوبة
            columns = columns_form.cleaned_data['columns']
            
            # تصدير التقرير بناءً على الصيغة المختارة
            if report.export_format == 'pdf':
                return export_to_pdf(projects, columns)
            elif report.export_format == 'excel':
                return export_to_excel(projects, columns)
            elif report.export_format == 'csv':
                return export_to_csv(projects, columns)
    else:
        form = ReportForm()
        columns_form = ExportColumnsForm()
    return render(request, 'management/create_report.html', {'form': form, 'columns_form': columns_form})

def export_to_excel(projects, columns):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="projects.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "المشاريع"

    # إضافة العناوين بالأعمدة المحددة
    ws.append([smart_str(col) for col in columns])

    # إضافة البيانات
    for project in projects:
        row = [smart_str(getattr(project, col)) for col in columns]
        ws.append(row)

    wb.save(response)
    return response

def export_to_pdf(projects, columns):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="projects.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []

    styles = getSampleStyleSheet()
    title = Paragraph(smart_str("قائمة المشاريع"), styles['Title'])
    elements.append(title)

    # إضافة العناوين بالأعمدة المحددة
    data = [[smart_str(col) for col in columns]]

    # إضافة البيانات
    for project in projects:
        row = [smart_str(getattr(project, col)) for col in columns]
        data.append(row)

    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))

    elements.append(table)
    doc.build(elements)

    return response

def export_to_csv(projects, columns):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="projects.csv"'
    response.write(u'\ufeff'.encode('utf8'))  # BOM لتجنب مشاكل الترميز

    writer = csv.writer(response)
    writer.writerow([smart_str(col) for col in columns])  # العناوين بالأعمدة المحددة

    for project in projects:
        row = [smart_str(getattr(project, col)) for col in columns]
        writer.writerow(row)

    return response

# عرض قائمة التقارير
def report_list(request):
    reports = Report.objects.all()
    return render(request, 'management/report_list.html', {'reports': reports})

# عرض تفاصيل تقرير معين
def report_detail(request, report_id):
    report = get_object_or_404(Report, id=report_id)
    return render(request, 'management/report_detail.html', {'report': report})

# عرض التقويم
def calendar_view(request):
    maintenance_tasks = MaintenanceTask.objects.all()
    return render(request, 'management/calendar.html', {'maintenance_tasks': maintenance_tasks})
